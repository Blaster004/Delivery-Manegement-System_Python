import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

class DeliveryManagementSystem:
    def __init__(self):
        self.excel_file = self.get_daily_excel_filename()
        self.deliveries = []
        self.load_from_file()

    def get_daily_excel_filename(self):
        today = datetime.now().strftime("%Y-%m-%d")
        return f"deliveries_{today}.xlsx"

    def load_from_file(self):
        try:
            df = pd.read_excel(self.excel_file, engine="openpyxl")
            if df.empty:
                print(f"{self.excel_file} is empty.")
            else:
                self.deliveries = df.to_dict(orient="records")
                print(f"Data loaded from {self.excel_file}.")
        except FileNotFoundError:
            print(f"File {self.excel_file} not found. Starting with an empty system.")
        except Exception as e:
            print(f"Error loading data from {self.excel_file}: {e}. Starting with an empty system.")

    def save_to_file(self):
        # Add "Total Price" to deliveries
        for delivery in self.deliveries:
            delivery["Total Price"] = delivery["Product Price"] + delivery["Delivery Fee"]

        df = pd.DataFrame(self.deliveries)

        with pd.ExcelWriter(self.excel_file, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Deliveries")
            workbook = writer.book
            worksheet = workbook["Deliveries"]

            # Format the columns
            for col in worksheet.columns:
                max_length = 0
                for cell in col:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                col_letter = get_column_letter(col[0].column)
                worksheet.column_dimensions[col_letter].width = max_length + 2

            worksheet.column_dimensions["A"].width = 21

            # Color the status column
            for row in worksheet.iter_rows(min_row=2, min_col=9, max_col=9):
                for cell in row:
                    if cell.value == "Pending":
                        cell.font = Font(color="FFFF00")
                    elif cell.value == "In Progress":
                        cell.font = Font(color="0000FF")
                    elif cell.value == "Delivered":
                        cell.font = Font(color="00FF00")
                    elif cell.value == "Cancelled":
                        cell.font = Font(color="FF0000")
                    else:
                        cell.font = Font(color="000000")

            # Calculate totals for summary
            total_sell = sum(order["Product Price"] for order in self.deliveries if order["Status"] == "Delivered")
            total_delivery_fee = sum(
                order["Delivery Fee"] for order in self.deliveries if order["Status"] == "Delivered")
            total_profit = sum(
                (order["Product Price"] - order["Buying Price"])  # Profit calculation
                for order in self.deliveries if order["Status"] == "Delivered"
            )

            summary_start_row = len(self.deliveries) + 4

            thin_border = Border(bottom=Side(style="thin"))
            for col in range(1, len(df.columns) + 1):
                worksheet.cell(row=summary_start_row - 1, column=col).border = thin_border

            summary_cell = worksheet.cell(row=summary_start_row, column=1, value="Summary")
            summary_cell.font = Font(bold=True, color="800080", size=16)
            worksheet.merge_cells(start_row=summary_start_row, start_column=1, end_row=summary_start_row, end_column=3)
            summary_cell.alignment = Alignment(horizontal="center", vertical="center")

            worksheet.cell(row=summary_start_row + 1, column=1, value="Total Sell").font = Font(bold=True)
            worksheet.cell(row=summary_start_row + 2, column=1, value="Total Delivery Fee").font = Font(bold=True)
            worksheet.cell(row=summary_start_row + 3, column=1, value="Total Profit").font = Font(bold=True)

            total_sell_cell = worksheet.cell(row=summary_start_row + 1, column=2, value=total_sell)
            total_sell_cell.font = Font(color="0000FF")
            total_sell_cell.alignment = Alignment(horizontal="center", vertical="center")

            total_delivery_fee_cell = worksheet.cell(row=summary_start_row + 2, column=2, value=total_delivery_fee)
            total_delivery_fee_cell.font = Font(color="FF1493")
            total_delivery_fee_cell.alignment = Alignment(horizontal="center", vertical="center")

            profit_cell = worksheet.cell(row=summary_start_row + 3, column=2, value=total_profit)
            profit_cell.font = Font(color="00FF00", bold=True)
            profit_cell.alignment = Alignment(horizontal="center", vertical="center")

        print(f"Data saved to {self.excel_file} with colored statuses, totals, and Total Price column.")

    def add_order(self, customer_name, phone_number, product_name, buying_price, product_price):
        print("\nSelect Delivery Location:")
        print("1. Inside Dhaka (60 Taka Delivery Charge)")
        print("2. Outside Dhaka (100 Taka Delivery Charge)")

        try:
            location_choice = int(input("Enter your choice (1 or 2): "))
            if location_choice == 1:
                delivery_fee = 60
            elif location_choice == 2:
                delivery_fee = 100
            else:
                print("Invalid choice. Defaulting to 60 Taka delivery charge (Inside Dhaka).")
                delivery_fee = 60
        except ValueError:
            print("Invalid input. Defaulting to 60 Taka delivery charge (Inside Dhaka).")
            delivery_fee = 60

        address = input("Enter delivery address: ")

        new_order = {
            "ID": len(self.deliveries) + 1,
            "Customer Name": customer_name,
            "Phone Number": phone_number,
            "Address": address,
            "Product Name": product_name,
            "Buying Price": buying_price,
            "Product Price": product_price,
            "Delivery Fee": delivery_fee,
            "Status": "Pending",
            "Total Price": product_price + delivery_fee  # Calculating Total Price here
        }
        self.deliveries.append(new_order)
        print("Order added successfully.")
        self.save_to_file()

    def update_status(self, order_id):
        order_found = False
        for delivery in self.deliveries:
            if delivery["ID"] == order_id:
                print("\nSelect a new status for the delivery:")
                print("1. Pending")
                print("2. In Progress")
                print("3. Delivered")
                print("4. Cancelled")
                try:
                    status_choice = int(input("Enter the number corresponding to the status: "))
                    if status_choice == 1:
                        new_status = "Pending"
                    elif status_choice == 2:
                        new_status = "In Progress"
                    elif status_choice == 3:
                        new_status = "Delivered"
                    elif status_choice == 4:
                        new_status = "Cancelled"
                    else:
                        print("Invalid choice. Status not updated.")
                        return
                except ValueError:
                    print("Invalid input. Please enter a number.")
                    return

                delivery["Status"] = new_status
                delivery["Total Price"] = delivery["Product Price"] + delivery["Delivery Fee"]  # Recalculate Total Price after status change
                print(f"Order ID {order_id} status updated to {new_status}.")
                self.save_to_file()
                order_found = True
                break

        if not order_found:
            print(f"Order ID {order_id} not found. Please check the order ID and try again.")

    def view_orders(self):
        if self.deliveries:
            print("\nAll Orders:")
            for order in self.deliveries:
                print(
                    f"ID: {order['ID']}, Customer: {order['Customer Name']}, Product: {order['Product Name']}, Status: {order['Status']}, Total Price: {order['Total Price']}")
        else:
            print("No orders available.")

    def delete_all_data(self):
        confirm = input("Are you sure you want to delete all data? This action cannot be undone. (y/n): ")
        if confirm.lower() == 'y':
            self.deliveries = []
            print("All data deleted.")
            self.save_to_file()

def main():
    system = DeliveryManagementSystem()

    while True:
        print("\nDelivery Management System")
        print("1. Add Delivery Order")
        print("2. Update Delivery Status")
        print("3. View All Orders")
        print("4. Delete All Data")
        print("5. Save and Exit")
        choice = input("Enter your choice: ")

        if choice == "1":
            customer_name = input("Enter customer name: ")
            phone_number = input("Enter phone number: ")
            product_name = input("Enter product name: ")
            try:
                product_price = float(input("Enter product price: "))
                buying_price = float(input("Enter buying price: "))
                system.add_order(customer_name, phone_number, product_name, buying_price, product_price)
            except ValueError:
                print("Invalid input. Please enter numerical values for price.")

        elif choice == "2":
            try:
                order_id = int(input("Enter order ID to update: "))
                system.update_status(order_id)
            except ValueError:
                print("Invalid order ID. Please enter a number.")

        elif choice == "3":
            system.view_orders()

        elif choice == "4":
            system.delete_all_data()

        elif choice == "5":
            print("Exiting the system. Goodbye!")
            break

        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
